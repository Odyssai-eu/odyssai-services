"""Document Producer — render model output into real Office files.

The mirror of the Docling parser: Docling turns .docx/.xlsx INTO markdown on
the way IN; this turns markdown / a JSON table-spec INTO real .docx/.xlsx on
the way OUT. Lives in the odyssai-services sidecar; the Companion add-on
(skill-triggered) POSTs here and attaches the returned file.

Two surfaces:
  * DOCX  — the model already emits markdown; pandoc renders it to a REAL Word
    document (not a markdown file). V1 = faithful body (headings/lists/tables/
    bold/code). V2 (noted, not built) = dynamic updatable TOC field + a
    house-style reference.docx.
  * XLSX  — the model emits a JSON spec (sheets / rows / header / formulas /
    number-formats / widths); openpyxl builds a genuine workbook.

Design choice: pandoc for docx (best-fidelity markdown->docx, single binary the
container installs), openpyxl for xlsx (multi-sheet, formulas, formats — what
separates a real workbook from a CSV in disguise).
"""

from __future__ import annotations

import io
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Optional

# ──────────────────────────────────────────────────────────────────────────────
# DOCX — markdown -> real Word document (pandoc)
# ──────────────────────────────────────────────────────────────────────────────

DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def pandoc_available() -> bool:
    return shutil.which("pandoc") is not None


def md_to_docx(markdown: str, *, reference_docx: Optional[str] = None,
               toc: bool = False) -> bytes:
    """Render markdown to a real .docx via pandoc. Returns the file bytes.

    reference_docx: path to a house-style template (pandoc --reference-doc).
    toc: pandoc's static table of contents. NOTE: this is a plain generated
         list, NOT a live Word TOC field — a truly *dynamic/updatable* TOC is
         V2 (python-docx post-processing to inject a TOC field). Kept as a flag
         so the API surface is stable when V2 lands.
    """
    if not pandoc_available():
        raise RuntimeError(
            "pandoc not installed — required for docx rendering "
            "(container: `apt-get install -y pandoc`)")
    if not markdown.strip():
        raise ValueError("empty markdown")

    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "out.docx"
        cmd = ["pandoc", "-f", "markdown+pipe_tables+grid_tables",
               "-t", "docx", "-o", str(out)]
        if toc:
            cmd += ["--toc", "--toc-depth=3"]
        if reference_docx and Path(reference_docx).exists():
            cmd += ["--reference-doc", reference_docx]
        proc = subprocess.run(cmd, input=markdown.encode("utf-8"),
                              capture_output=True, timeout=120)
        if proc.returncode != 0:
            raise RuntimeError(
                f"pandoc failed ({proc.returncode}): "
                f"{proc.stderr.decode('utf-8', 'replace')[:300]}")
        return out.read_bytes()


# ──────────────────────────────────────────────────────────────────────────────
# XLSX — JSON spec -> real workbook (openpyxl)
# ──────────────────────────────────────────────────────────────────────────────
#
# Spec schema (all keys optional except sheets[].rows):
# {
#   "sheets": [
#     {
#       "name": "Q4",                         # sheet title (default Sheet1...)
#       "rows": [["Produit","CA"],["A",1200]],# 2D array; strings/numbers/None
#       "header": true,                       # bold + fill + freeze first row
#       "formulas": {"B4": "=SUM(B2:B3)"},    # cell -> formula string
#       "number_formats": {"B": "#,##0"},     # column letter -> excel format
#       "widths": {"A": 24, "B": 14},         # column letter -> width
#       "autofit": true                       # rough auto-width from content
#     }
#   ]
# }

def _apply_header(ws, ncols: int) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def spec_to_xlsx(spec: dict[str, Any]) -> bytes:
    """Build a real .xlsx workbook from a JSON spec. Returns file bytes."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string

    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("spec.sheets must be a non-empty list")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet; we add our own

    for i, sh in enumerate(sheets):
        rows = sh.get("rows")
        if not isinstance(rows, list):
            raise ValueError(f"sheet[{i}].rows must be a list of lists")
        name = str(sh.get("name") or f"Sheet{i + 1}")[:31]  # Excel 31-char cap
        ws = wb.create_sheet(title=name)

        ncols = 0
        for r, row in enumerate(rows, start=1):
            cells = row if isinstance(row, list) else [row]
            ncols = max(ncols, len(cells))
            for c, val in enumerate(cells, start=1):
                ws.cell(row=r, column=c, value=val)

        # Formulas override cell values (written as "=...").
        for ref, formula in (sh.get("formulas") or {}).items():
            ws[ref] = formula if str(formula).startswith("=") else f"={formula}"

        if sh.get("header") and rows:
            _apply_header(ws, ncols)

        for col, fmt in (sh.get("number_formats") or {}).items():
            idx = column_index_from_string(col)
            for r in range(1, len(rows) + 1):
                ws.cell(row=r, column=idx).number_format = fmt

        widths = sh.get("widths") or {}
        for col, w in widths.items():
            ws.column_dimensions[col].width = float(w)
        if sh.get("autofit"):
            for c in range(1, ncols + 1):
                letter = get_column_letter(c)
                if letter in widths:
                    continue
                longest = max(
                    (len(str(ws.cell(row=r, column=c).value or ""))
                     for r in range(1, len(rows) + 1)), default=8)
                ws.column_dimensions[letter].width = min(max(longest + 2, 8), 60)

    if not wb.sheetnames:
        raise ValueError("no sheets produced")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# FastAPI router (mounted by odyssai-services api.py)
# ──────────────────────────────────────────────────────────────────────────────

try:
    from pydantic import BaseModel

    class DocxReq(BaseModel):
        markdown: str
        filename: str = "document.docx"
        toc: bool = False

    class XlsxReq(BaseModel):
        spec: dict
        filename: str = "workbook.xlsx"
except Exception:  # pydantic absent in a non-serving env
    DocxReq = XlsxReq = None  # type: ignore


def build_router():
    # Request models MUST live at MODULE level (above), not inside this
    # closure: FastAPI resolves endpoint type-hints via the module globals,
    # and a class defined in the closure is invisible there -> the param is
    # mistaken for a query field (HTTP 422 "req required in query"). Learned
    # the hard way on the first live deploy.
    from fastapi import APIRouter, HTTPException
    from fastapi.responses import Response

    router = APIRouter(prefix="/render", tags=["render"])

    @router.get("/health")
    def health():
        return {"ok": True, "pandoc": pandoc_available(), "xlsx": True}

    @router.post("/docx")
    def render_docx(req: DocxReq):
        try:
            data = md_to_docx(req.markdown, toc=req.toc)
        except Exception as e:
            raise HTTPException(400, str(e))
        return Response(content=data, media_type=DOCX_MIME, headers={
            "Content-Disposition": f'attachment; filename="{req.filename}"'})

    @router.post("/xlsx")
    def render_xlsx(req: XlsxReq):
        try:
            data = spec_to_xlsx(req.spec)
        except Exception as e:
            raise HTTPException(400, str(e))
        return Response(content=data, media_type=XLSX_MIME, headers={
            "Content-Disposition": f'attachment; filename="{req.filename}"'})

    return router


if __name__ == "__main__":
    # Self-test: produce real files to /tmp and validate them.
    import sys
    md = ("# Rapport Q4\n\n## Résumé\n\nLe chiffre d'affaires a **progressé** "
          "de 12%.\n\n- Produit A : forte croissance\n- Produit B : stable\n\n"
          "| Produit | CA | Marge |\n|---|---|---|\n| A | 1200 | 18% |\n"
          "| B | 3400 | 22% |\n\n> Objectif dépassé.\n")
    spec = {"sheets": [{
        "name": "CA Q4",
        "rows": [["Produit", "CA", "Coût"], ["A", 1200, 984],
                 ["B", 3400, 2652], ["Total", None, None]],
        "header": True,
        "formulas": {"B4": "=SUM(B2:B3)", "C4": "=SUM(C2:C3)"},
        "number_formats": {"B": "#,##0", "C": "#,##0"},
        "autofit": True,
    }]}

    xlsx = spec_to_xlsx(spec)
    Path("/tmp/render_test.xlsx").write_bytes(xlsx)
    print(f"xlsx: {len(xlsx)} bytes -> /tmp/render_test.xlsx")

    if pandoc_available():
        docx = md_to_docx(md)
        Path("/tmp/render_test.docx").write_bytes(docx)
        print(f"docx: {len(docx)} bytes -> /tmp/render_test.docx")
    else:
        print("pandoc absent — skipping docx (install: brew install pandoc)")
        sys.exit(0)
